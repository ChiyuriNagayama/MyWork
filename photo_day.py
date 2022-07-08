#!/usr/bin/env python
# coding: utf-8

# In[1]:


import random
import openpyxl
import datetime

def tag():
    tag = ["work", "experiment", "room", "home", "walking", "science", "lifestyle", "city", "park",
          "color", "nature", "dynamics", "sea", "summer", "neon", "urban", "fireworks", "light", "flower",
          "forest", "spectral", "world", "architecture", "rainbow", "tulip"]
    return random.choice(tag)
title = tag()

print(title + "が今回のキーワードです。")

print("pixabayを立ち上げてください。")

url1_list = []

while True:
    url1_path = input("画像ファイルパスを入力してください。")
    
    if url1_path == "":
        break
    
    url1_list.append(url1_path)

print("unsplashを立ち上げてください。")

url2_list = []

while True:
    url2_path = input("画像ファイルパスを入力してください。")
    
    if url2_path == "":
        break
    
    url2_list.append(url2_path)
    
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = title

sheet.cell(1, 1, value = "pixabay")

for i in range(0, len(url1_list)):
    sheet.cell(i + 2, 1, value = url1_list[i])
    
sheet.cell(len(url1_list) + 3, 1, value = "unsplash")
    
for i in range(0, len(url2_list)):
    sheet.cell(len(url1_list) + i + 4, 1, value = url2_list[i])
    
d = datetime.datetime.today()
ymd = d.strftime("%Y%m%d")
    
wb.save(ymd + "_画像データ.xlsx")


# In[ ]:




