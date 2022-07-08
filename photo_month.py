#!/usr/bin/env python
# coding: utf-8

# In[3]:


import random
import openpyxl
import datetime

key = ["work", "experiment", "room", "home", "walking", "science", "lifestyle", "city", "park",
      "color", "nature", "dynamics", "sea", "summer", "neon", "urban", "fireworks", "light", "flower",
      "forest", "spectral", "world", "architecture", "rainbow", "tulip"]

rand_num = random.sample(range(0, len(key)), 5)

wb = openpyxl.Workbook()
sheet = wb.active

d = datetime.datetime.today()
ymd = d.strftime("%Y%m%d")

print("pixabayを立ち上げてください。")

k = 1

for i in range(0, 5):
    l = rand_num[i]
    print("pixabayで、" + str(key[l]) + "を検索してください。")

    url1_list = []

    while True:
        url1_path = input("画像ファイルパスを入力してください。")

        if url1_path == "":
            break

        url1_list.append(url1_path)
        
    sheet.cell(k, 1, value = "pixabay：" + str(key[l]))
    k += 1

    for j in range(0, len(url1_list)):
        sheet.cell(k, 1, value = url1_list[j])
        k += 1

print("unsplashを立ち上げてください。")

for i in range(0, 5):
    l = rand_num[i]
    print("unsplashで、" + str(key[l]) + "を検索してください。")

    url2_list = []

    while True:
        url2_path = input("画像ファイルパスを入力してください。")

        if url2_path == "":
            break

        url2_list.append(url2_path)

    sheet.cell(k, 1, value = "unsplash："+ str(key[l]))
    k += 1

    for j in range(0, len(url2_list)):
        sheet.cell(k, 1, value = url2_list[j])
        k += 1
    
wb.save(ymd + "_画像データ.xlsx")    


# In[ ]:




